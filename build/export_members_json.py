"""Build dashboards/data/members.json from the 119th Master List.

The sheet has messy rows: seats that changed hands during the 119th carry a
second "Left during 119th (BIO)" row whose Bioguide-ID column is often a stray
number or a duplicate, and sometimes blank names. We key each member by a
*canonical* Bioguide and keep the best (most complete, currently-sitting) row.
"""
import json
import re
from pathlib import Path

import openpyxl

SRC = "/Users/amandakoski/Documents/Claude/Suffrage and Sass/Sovereign AI/data/master/119th_MASTER_LIST.xlsx"
DEST = Path("/Users/amandakoski/Documents/Claude/Suffrage and Sass/Sovereign AI/dashboards/data/members.json")

BIO_RE = re.compile(r"^[A-Z]\d{6}$")

wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))
hdr = [str(h).strip() if h is not None else "" for h in rows[0]]
idx = {h: i for i, h in enumerate(hdr)}


def cell(row, name):
    i = idx.get(name)
    if i is None:
        return ""
    v = row[i]
    if v is None:
        return ""
    return str(v).strip()


def split_list(s):
    return [p.strip() for p in re.split(r"[;\n]", s) if p.strip()]


def canonical_bio(row):
    col0 = cell(row, "Bioguide ID")
    status = cell(row, "Status / Bioguide")
    if BIO_RE.match(col0):
        return col0
    m = re.search(r"\(([A-Z]\d{6})\)", status)
    return m.group(1) if m else None


def completeness(row):
    return sum(1 for v in row if v not in (None, ""))


def build_record(row):
    rec = {
        "bioguide": canonical_bio(row),
        "full_name": cell(row, "Full Name"),
        "first_name": cell(row, "First Name"),
        "last_name": cell(row, "Last Name"),
        "state": cell(row, "State"),
        "district": cell(row, "District"),
        "party": cell(row, "Party"),
        "gender": cell(row, "Gender"),
        "birthday": cell(row, "Birthday"),
        "phone": cell(row, "Phone"),
        "office": cell(row, "Office"),
        "year_elected": cell(row, "Year Elected to Current Seat"),
        "term_ends": cell(row, "Year Current Term Ends"),
        "term_start": cell(row, "Term Start"),
        "term_end": cell(row, "Term End"),
        "gov_website": cell(row, "Official Government Website"),
        "campaign_website": cell(row, "Campaign Website"),
        "twitter_official": cell(row, "Official Twitter"),
        "twitter_personal": cell(row, "Campaign/Personal Twitter"),
        "instagram_official": cell(row, "Official Instagram"),
        "instagram_personal": cell(row, "Campaign/Personal Instagram"),
        "facebook": cell(row, "Facebook"),
        "youtube": cell(row, "YouTube"),
        "committees": split_list(cell(row, "Committees")),
        "fd_recent": cell(row, "Most Recent House Financial Disclosure"),
        "fd_others": split_list(cell(row, "Other Financial Disclosure Docs")),
        "fec_house": cell(row, "FEC ID - House Campaign"),
        "fec_leadership": cell(row, "FEC ID - Leadership PAC"),
        "fec_other_pac": cell(row, "FEC ID - Other PAC"),
        "fec_senate": cell(row, "FEC ID - Senate Run"),
        "fec_president": cell(row, "FEC ID - President Run"),
        "ran_state_office": cell(row, "Ran for State-Level Office (Y/N)"),
        "state_office": cell(row, "Which State Office"),
        "state_office_state": cell(row, "Which State"),
        "status": cell(row, "Status / Bioguide"),
    }
    # drop empty values to keep the file lean; JS treats missing as empty
    return {k: v for k, v in rec.items() if v not in ("", [], None)}


best = {}  # bio -> (is_sitting, completeness, record)
for row in rows[1:]:
    bio = canonical_bio(row)
    if not bio:
        continue
    rec = build_record(row)
    sitting = cell(row, "Status / Bioguide").lower().startswith("sitting")
    score = (1 if sitting else 0, completeness(row))
    if bio not in best or score > best[bio][0]:
        best[bio] = (score, rec)

members = {bio: rec for bio, (score, rec) in best.items()}
result = {
    "generated_from": "119th_MASTER_LIST.xlsx",
    "count": len(members),
    "members": members,
}
DEST.parent.mkdir(parents=True, exist_ok=True)
DEST.write_text(json.dumps(result, indent=1))
print(f"Wrote {len(members)} members → {DEST} ({DEST.stat().st_size:,} bytes)")

# sanity: the dashboard's loaded members
for b in ["B001324", "C001067", "D000594", "D000399", "F000484",
          "K000391", "L000598", "N000026", "R000610", "S001193"]:
    r = members.get(b, {})
    print(f"  {b}: {r.get('full_name','<MISSING>')!r} "
          f"committees={len(r.get('committees',[]))} "
          f"socials={sum(1 for k in ('twitter_official','twitter_personal','instagram_official','instagram_personal','facebook','youtube') if r.get(k))} "
          f"phone={'Y' if r.get('phone') else '-'}")
